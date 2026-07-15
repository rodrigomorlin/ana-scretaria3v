"""
Ana v3 — Secretária Virtual de Anestesiologia
PostgreSQL + SQLite, email, Google Calendar, relatórios
"""

from fastapi import FastAPI, HTTPException, Depends, Request, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, Any, List
import os, logging, hashlib, secrets, json, base64, io
import re, time
from datetime import datetime, timedelta
import urllib.request, urllib.error, urllib.parse

try:
    from pypdf import PdfReader
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(levelname)s | %(message)s')
log = logging.getLogger("ana")

app = FastAPI(title="Ana v3", version="3.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

SECRET         = os.environ.get("SECRET_KEY", "ana-secretaria-default-secret-change-me")
GROQ_KEY       = os.environ.get("GROQ_API_KEY", "")
CEREBRAS_API_KEY = os.environ.get("CEREBRAS_API_KEY", "")
CEREBRAS_MODEL   = os.environ.get("CEREBRAS_MODEL", "llama-3.3-70b")
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
EMAIL_FROM     = os.environ.get("EMAIL_FROM", "A.N.A <onboarding@resend.dev>")
SMTP_HOST      = os.environ.get("SMTP_HOST", "")
SMTP_PORT      = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER      = os.environ.get("SMTP_USER", "")
SMTP_PASS      = os.environ.get("SMTP_PASS", "")
SMTP_FROM      = os.environ.get("SMTP_FROM", "ana@grupo-anestesia.com")
GCAL_CREDS     = os.environ.get("GCAL_CREDENTIALS", "")
GCAL_ID        = os.environ.get("GCAL_CALENDAR_ID", "primary")
GOOGLE_CLIENT_ID     = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
APP_BASE_URL         = os.environ.get("APP_BASE_URL", "")
GOOGLE_ROUTES_API_KEY = os.environ.get("GOOGLE_ROUTES_API_KEY", "")
GEMINI_API_KEY       = os.environ.get("GEMINI_API_KEY", "")
OCR_SPACE_API_KEY    = os.environ.get("OCR_SPACE_API_KEY", "")
VAPID_PUBLIC_KEY     = os.environ.get("VAPID_PUBLIC_KEY", "")
VAPID_PRIVATE_KEY    = os.environ.get("VAPID_PRIVATE_KEY", "")
VAPID_CLAIMS_EMAIL   = os.environ.get("VAPID_CLAIMS_EMAIL", "rodrigomorlin@gmail.com")

# ── SUPABASE (migração em andamento — coexiste com auth antigo) ──
SUPABASE_URL              = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_ANON_KEY         = os.environ.get("SUPABASE_ANON_KEY", "")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
if not (SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY and SUPABASE_ANON_KEY):
    raise RuntimeError("Configure SUPABASE_URL, SUPABASE_ANON_KEY e SUPABASE_SERVICE_ROLE_KEY no Railway.")
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import data as ana_data

# ── AUTH ───────────────────────────────────────────────────
def auth(request: Request):
    """Autenticação: JWT do Supabase (Authorization: Bearer <token>)."""
    authz = request.headers.get("Authorization", "")
    if authz.startswith("Bearer ") and SUPABASE_URL:
        user = _auth_supabase(request, authz[7:])
        if user:
            return user
    raise HTTPException(401, "Não autorizado.")

def db_log(nivel, msg, usuario="", ip="", org_id=""):
    """Log de aplicação (o log persistente por grupo é o ana_logs, via data.py)."""
    log.info(f"[{nivel}] {msg}" + (f" · user={usuario}" if usuario else ""))

# ── SUPABASE AUTH (JWT via JWKS) ─────────────────────────────
_jwks_client = None
_membership_cache: dict = {}  # user_id → (timestamp, [memberships])

def _get_jwks_client():
    global _jwks_client
    if _jwks_client is None:
        import jwt as pyjwt
        _jwks_client = pyjwt.PyJWKClient(
            f"{SUPABASE_URL}/auth/v1/.well-known/jwks.json",
            cache_keys=True, lifespan=3600)
    return _jwks_client

def validate_supabase_jwt(token: str) -> Optional[dict]:
    """Valida um JWT do Supabase. Retorna claims ou None."""
    try:
        import jwt as pyjwt
        header = pyjwt.get_unverified_header(token)
        alg = header.get("alg", "")
        if alg in ("RS256", "ES256"):
            key = _get_jwks_client().get_signing_key_from_jwt(token).key
            claims = pyjwt.decode(token, key, algorithms=[alg], audience="authenticated")
        elif alg == "HS256":
            # Projetos antigos do Supabase assinam com o JWT secret (não exposto);
            # sem o secret não dá para validar HS256 — rejeita com log claro.
            log.warning("JWT Supabase HS256 recebido — configure JWT assimétrico no projeto ou forneça o secret.")
            return None
        else:
            return None
        return claims
    except Exception as e:
        log.warning(f"JWT Supabase inválido: {type(e).__name__}: {e}")
        return None

def sb_rest(method: str, path: str, body=None, use_service_role=True, user_jwt: str = ""):
    """Chamada ao PostgREST do Supabase. path ex: '/group_members?user_id=eq.xxx&select=group_id,role'"""
    key = SUPABASE_SERVICE_ROLE_KEY if use_service_role else SUPABASE_ANON_KEY
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {user_jwt or key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(f"{SUPABASE_URL}/rest/v1{path}", data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            raw = r.read()
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        body_err = e.read().decode("utf-8", errors="ignore")[:300]
        log.error(f"Supabase REST {method} {path} → {e.code}: {body_err}")
        raise HTTPException(502, f"Erro Supabase ({e.code})")

def get_memberships(user_id: str) -> list:
    """Retorna [{group_id, role}] do usuário, com cache de 60s."""
    import time as _t
    cached = _membership_cache.get(user_id)
    if cached and _t.time() - cached[0] < 60:
        return cached[1]
    rows = sb_rest("GET", f"/group_members?user_id=eq.{user_id}&select=group_id,role")
    _membership_cache[user_id] = (_t.time(), rows)
    return rows

def _auth_supabase(request: Request, token: str) -> Optional[dict]:
    """Autentica via JWT Supabase. Retorna user dict compatível com o resto do código."""
    claims = validate_supabase_jwt(token)
    if not claims:
        return None
    user_id = claims.get("sub", "")
    email = claims.get("email", "")
    if not user_id:
        return None
    memberships = get_memberships(user_id)
    if not memberships:
        raise HTTPException(403, "Usuário não pertence a nenhum grupo. Peça convite ao administrador.")
    # Multi-grupo: header X-Group-Id seleciona; valida pertencimento
    wanted = request.headers.get("X-Group-Id", "")
    if wanted:
        m = next((m for m in memberships if m["group_id"] == wanted), None)
        if not m:
            raise HTTPException(403, "Você não pertence a esse grupo.")
    else:
        m = memberships[0]
    # Nome: metadados do JWT ou email
    meta = claims.get("user_metadata") or {}
    nome = meta.get("full_name") or meta.get("name") or (email.split("@")[0] if email else user_id[:8])
    role = "admin" if m.get("role") == "admin" else "medico"
    return {"id": user_id, "nome": nome, "role": role,
            "org_id": m["group_id"], "email": email,
            "auth_source": "supabase", "jwt": token,
            "memberships": memberships}

# ── CONFIG PERSISTENTE (chave/valor no banco, por organização) ──
def get_config(chave: str, default: str = "", org_id: str = "") -> str:
    """Config por grupo — ana_gcal_config no Supabase (calendar_id)."""
    if chave != "gcal_calendar_id" or not org_id:
        return default
    try:
        rows = sb_rest("GET", f"/ana_gcal_config?group_id=eq.{org_id}&select=calendar_id")
        return (rows[0].get("calendar_id") or default) if rows else default
    except Exception:
        return default

def set_config(chave: str, valor: str, org_id: str = ""):
    if chave != "gcal_calendar_id" or not org_id:
        return
    rows = sb_rest("GET", f"/ana_gcal_config?group_id=eq.{org_id}&select=group_id")
    if rows:
        sb_rest("PATCH", f"/ana_gcal_config?group_id=eq.{org_id}", {"calendar_id": valor, "updated_at": "now()"})
    else:
        sb_rest("POST", "/ana_gcal_config", {"group_id": org_id, "calendar_id": valor})

def get_gcal_id(org_id: str = "default") -> str:
    """Prioriza o valor configurado pelo usuário no app; cai para a env var como fallback."""
    return get_config("gcal_calendar_id", GCAL_ID, org_id=org_id)

# ── EMAIL ──────────────────────────────────────────────────
async def send_email(to, subject, body):
    """Envio de email: Resend (RESEND_API_KEY) com fallback SMTP."""
    if not to: return
    if RESEND_API_KEY:
        try:
            payload = json.dumps({"from": EMAIL_FROM, "to": [to],
                                  "subject": subject, "html": body}).encode("utf-8")
            http_req = urllib.request.Request(
                "https://api.resend.com/emails", data=payload, method="POST",
                headers={"Content-Type": "application/json",
                         "Authorization": f"Bearer {RESEND_API_KEY}"})
            with urllib.request.urlopen(http_req, timeout=20) as resp:
                r = json.loads(resp.read())
            log.info(f"Email (Resend) → {to} · id={r.get('id','?')}")
            return
        except urllib.error.HTTPError as e:
            log.error(f"Resend erro {e.code}: {e.read().decode('utf-8','ignore')[:300]}")
            return
        except Exception as e:
            log.error(f"Resend erro: {e}")
            return
    if not SMTP_HOST: return
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject; msg["From"] = SMTP_FROM; msg["To"] = to
        msg.attach(MIMEText(body, "html", "utf-8"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls(); s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(SMTP_FROM, to, msg.as_string())
        log.info(f"Email (SMTP) → {to}")
    except Exception as e: log.error(f"Email erro: {e}")

def email_html(ev, setor_name):
    return f"""<div style="font-family:Arial;max-width:480px;margin:0 auto;padding:20px">
      <div style="background:#6C63D4;color:#fff;border-radius:10px 10px 0 0;padding:14px 18px">
        <b>A.N.A · Novo agendamento</b></div>
      <div style="background:#f9f9ff;border:1px solid #E4E4EF;border-radius:0 0 10px 10px;padding:18px">
        <table style="width:100%;font-size:13px">
          <tr><td style="color:#888;padding:4px 0;width:110px">Setor</td><td><b>{setor_name}</b></td></tr>
          <tr><td style="color:#888;padding:4px 0">Procedimento</td><td>{ev.get('proc','—')}</td></tr>
          <tr><td style="color:#888;padding:4px 0">Paciente</td><td>{ev.get('paciente','—')}</td></tr>
          <tr><td style="color:#888;padding:4px 0">Data</td><td>{ev.get('date','—')}</td></tr>
          <tr><td style="color:#888;padding:4px 0">Horário</td><td>{ev.get('time','—')}</td></tr>
          {f"<tr><td style='color:#888;padding:4px 0'>Obs</td><td>{ev.get('obs')}</td></tr>" if ev.get('obs') else ''}
        </table>
        <p style="font-size:10px;color:#aaa;margin-top:14px">A.N.A · Secretária Virtual</p>
      </div></div>"""

# ── HELPERS DE HORÁRIO ──────────────────────────────────────
def _time_to_min(t: str) -> int:
    """Converte 'HH:MM' em minutos desde meia-noite."""
    try:
        h, m = t.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return 0

def fmtDate(d: str) -> str:
    """Formata YYYY-MM-DD para DD/MM."""
    try: return f"{d[8:10]}/{d[5:7]}"
    except: return d

# ── DESLOCAMENTO ENTRE SETORES (HERE API + cache) ──────────
def _cache_key(setor_a: str, setor_b: str, org_id: str) -> str:
    return f"{org_id}:{setor_a}:{setor_b}"

def _google_routes_duration(origem: str, destino: str) -> Optional[int]:
    """Chama a Google Routes API v2 e retorna tempo em minutos com trânsito histórico.
    Usa TRAFFIC_AWARE_OPTIMAL para considerar padrões históricos de tráfego."""
    try:
        # Google Routes API v2 — endpoint único, sem necessidade de geocodificação prévia
        # aceita endereços diretamente no campo address
        payload = json.dumps({
            "origin": {"address": origem},
            "destination": {"address": destino},
            "travelMode": "DRIVE",
            "routingPreference": "TRAFFIC_AWARE_OPTIMAL",
            "departureTime": datetime.now().strftime("%Y-%m-%dT08:00:00Z"),
            "computeAlternativeRoutes": False,
            "routeModifiers": {"avoidTolls": False, "avoidHighways": False},
            "languageCode": "pt-BR",
            "units": "METRIC"
        }).encode("utf-8")

        req = urllib.request.Request(
            "https://routes.googleapis.com/directions/v2:computeRoutes",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "X-Goog-Api-Key": GOOGLE_ROUTES_API_KEY,
                "X-Goog-FieldMask": "routes.duration,routes.distanceMeters,routes.staticDuration",
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read())

        routes = data.get("routes", [])
        if not routes:
            log.warning(f"Google Routes: nenhuma rota encontrada de '{origem}' para '{destino}'")
            return None

        # duration considera trânsito; staticDuration ignora — usamos duration
        duration_str = routes[0].get("duration", "")
        if not duration_str:
            return None
        # formato "NNNs" (segundos) ex: "1234s"
        segundos = int(duration_str.rstrip("s"))
        minutos = max(1, round(segundos / 60))
        dist_km = routes[0].get("distanceMeters", 0) / 1000
        log.info(f"Google Routes: {origem} → {destino} = {minutos} min / {dist_km:.1f}km (com trânsito)")
        return minutos
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        log.error(f"Google Routes API erro {e.code}: {body[:300]}")
        return None
    except Exception as e:
        log.error(f"Google Routes API erro: {e}")
        return None

# ── ROTAS DE DESLOCAMENTO ───────────────────────────────────
@app.get("/api/deslocamentos")
def list_deslocamentos(user=Depends(auth)):
    return ana_data.sb_list_deslocamentos(user)

@app.post("/api/deslocamentos/recalcular")
def recalcular_deslocamentos(user=Depends(auth)):
    return ana_data.sb_recalcular_deslocamentos(user)

# ── GOOGLE CALENDAR ────────────────────────────────────────
GOOGLE_OAUTH_SCOPES = "https://www.googleapis.com/auth/calendar email"

def _oauth_redirect_uri() -> str:
    base = APP_BASE_URL.rstrip("/") if APP_BASE_URL else ""
    return f"{base}/api/oauth/google/callback"

def build_oauth_url(state: str) -> str:
    params = urllib.parse.urlencode({
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": _oauth_redirect_uri(),
        "response_type": "code",
        "scope": GOOGLE_OAUTH_SCOPES,
        "access_type": "offline",
        "prompt": "consent",
        "state": state,
    })
    return f"https://accounts.google.com/o/oauth2/v2/auth?{params}"

def exchange_oauth_code(code: str) -> dict:
    """Troca o código de autorização por access_token + refresh_token."""
    data = urllib.parse.urlencode({
        "code": code,
        "client_id": GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "redirect_uri": _oauth_redirect_uri(),
        "grant_type": "authorization_code",
    }).encode("utf-8")
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data, method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())

def refresh_oauth_token(refresh_token: str) -> dict:
    data = urllib.parse.urlencode({
        "refresh_token": refresh_token,
        "client_id": GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "grant_type": "refresh_token",
    }).encode("utf-8")
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data, method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())

def _gtok_get(uid: str) -> dict:
    """Tokens Google do usuário — Supabase (ana_user_google_tokens) ou SQLite."""
    try:
        rows = sb_rest("GET", f"/ana_user_google_tokens?user_id=eq.{uid}&select=credentials")
        return rows[0]["credentials"] if rows else {}
    except Exception:
        return {}

def _gtok_save(uid: str, **patch):
    cur = _gtok_get(uid)
    cur.update({k: v for k, v in patch.items() if v is not None})
    try:
        sb_rest("DELETE", f"/ana_user_google_tokens?user_id=eq.{uid}")
        sb_rest("POST", "/ana_user_google_tokens", {"user_id": uid, "credentials": cur})
    except Exception as e:
        log.error(f"_gtok_save supabase: {e}")
    return

def _gtok_clear(uid: str):
    try: sb_rest("DELETE", f"/ana_user_google_tokens?user_id=eq.{uid}")
    except Exception: pass
    return

def get_user_google_token(usuario_id: str) -> Optional[str]:
    """Retorna um access_token válido para o usuário, renovando se necessário. None se não conectado."""
    row = _gtok_get(usuario_id)
    if not row or not row.get("refresh_token"):
        return None
    expiry = row.get("expiry") or ""
    try:
        expiry_dt = datetime.fromisoformat(expiry) if expiry else datetime.min
    except Exception:
        expiry_dt = datetime.min
    if row.get("access_token") and datetime.now() < expiry_dt:
        return row["access_token"]
    # token expirado — renova
    log.info(f"GCal token: access_token expirado/ausente para {usuario_id}, renovando via refresh_token...")
    try:
        tok = refresh_oauth_token(row["refresh_token"])
        new_access = tok.get("access_token", "")
        if not new_access:
            log.error(f"GCal token: refresh não retornou access_token. Resposta: {tok}")
            return None
        expires_in = tok.get("expires_in", 3600)
        new_expiry = (datetime.now() + timedelta(seconds=expires_in - 60)).isoformat()
        _gtok_save(usuario_id, access_token=new_access, expiry=new_expiry)
        return new_access
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        log.error(f"GCal token: erro HTTP ao renovar token de {usuario_id}: {e.code} {body[:300]}")
        return None
    except Exception as e:
        log.error(f"Erro ao renovar token Google de {usuario_id}: {e}")
        return None

def _build_gcal_service(user_access_token: Optional[str] = None):
    """Retorna um serviço do Calendar API, usando OAuth do usuário se disponível, senão a conta de serviço."""
    from googleapiclient.discovery import build
    if user_access_token:
        import google.oauth2.credentials
        creds = google.oauth2.credentials.Credentials(token=user_access_token)
        return build("calendar", "v3", credentials=creds)
    if GCAL_CREDS:
        from google.oauth2.service_account import Credentials
        creds = Credentials.from_service_account_info(
            json.loads(GCAL_CREDS), scopes=["https://www.googleapis.com/auth/calendar"])
        return build("calendar", "v3", credentials=creds)
    return None

async def gcal_create(ev, setor_name, criado_por_id: Optional[str] = None, org_id: str = "default", event_id: str = ""):
    # Prioriza o calendário pessoal do médico que criou o evento (OAuth), senão usa a conta de serviço/calendário do grupo
    user_token = get_user_google_token(criado_por_id) if criado_por_id else None
    calendar_id = "primary" if user_token else get_gcal_id(org_id)
    log.info(f"GCal criar: criado_por_id={criado_por_id}, org_id={org_id}, tem_token_pessoal={bool(user_token)}, calendar_id={calendar_id}")
    try:
        svc = _build_gcal_service(user_token)
        if not svc:
            log.error("GCal criar: serviço não pôde ser construído (sem token e sem GCAL_CREDS).")
            return ""
        ini_min = _time_to_min(ev["time"])
        # Google Calendar exige hora de fim — usa 30min como padrão fixo (não afeta lógica de conflito)
        fim_min = min(ini_min + 30, 23*60+59)
        end_h, end_m = fim_min // 60, fim_min % 60
        body = {
            "summary": f"{ev['proc']} — {ev.get('paciente','—')}",
            "description": f"Médico: {ev['doc']}\nSetor: {setor_name}\nObs: {ev.get('obs','')}",
            "start": {"dateTime": f"{ev['date']}T{ev['time']}:00", "timeZone": "America/Sao_Paulo"},
            "end": {"dateTime": f"{ev['date']}T{end_h:02d}:{end_m:02d}:00", "timeZone": "America/Sao_Paulo"},
        }
        if event_id:
            # id determinístico (uuid do agendamento sem hífens) + etiqueta anaId para busca no cancelamento
            body["id"] = event_id
            body["extendedProperties"] = {"private": {"anaId": event_id}}
        try:
            r = svc.events().insert(calendarId=calendar_id, body=body).execute()
        except Exception as ins_err:
            # se o Google rejeitar o id customizado (400/409), recria sem id — a etiqueta anaId garante o cancelamento
            if event_id and any(t in str(ins_err) for t in ("400", "409", "identifier", "Invalid")):
                log.warning(f"GCal: id customizado rejeitado ({ins_err}) — recriando com etiqueta anaId")
                body.pop("id", None)
                r = svc.events().insert(calendarId=calendar_id, body=body).execute()
            else:
                raise
        log.info(f"GCal evento criado: {r.get('id')} (calendário: {'pessoal' if user_token else 'grupo'}, id={calendar_id})")
        return r.get("id","")
    except Exception as e:
        log.error(f"GCal criar — FALHA: {type(e).__name__}: {e}")
        return ""

async def gcal_delete(gcal_id, criado_por_id: Optional[str] = None, org_id: str = "default"):
    if not gcal_id: return
    user_token = get_user_google_token(criado_por_id) if criado_por_id else None
    calendar_id = "primary" if user_token else get_gcal_id(org_id)
    log.info(f"GCal delete: id={gcal_id}, calendário={'pessoal' if user_token else 'grupo'} ({calendar_id})")
    svc = _build_gcal_service(user_token)
    if not svc:
        log.warning("GCal delete: serviço indisponível (sem token e sem GCAL_CREDS)")
        return
    # 1º: delete direto pelo id determinístico
    try:
        svc.events().delete(calendarId=calendar_id, eventId=gcal_id).execute()
        log.info("GCal delete: removido pelo id ✓")
        return
    except Exception as e:
        log.warning(f"GCal delete por id falhou ({e}) — tentando busca pela etiqueta anaId")
    # 2º: fallback — busca pela etiqueta anaId (cobre eventos recriados sem id custom)
    try:
        res = svc.events().list(calendarId=calendar_id,
                                privateExtendedProperty=f"anaId={gcal_id}",
                                maxResults=5).execute()
        items = res.get("items", [])
        if not items:
            log.warning("GCal delete: nenhum evento com a etiqueta — provavelmente criado antes da versão atual (remover manualmente)")
            return
        for it in items:
            svc.events().delete(calendarId=calendar_id, eventId=it["id"]).execute()
        log.info(f"GCal delete: {len(items)} evento(s) removido(s) pela etiqueta ✓")
    except Exception as e:
        log.error(f"GCal delete (fallback etiqueta): {e}")

# ── MODELOS ────────────────────────────────────────────────
class LoginData(BaseModel):
    usuario_id: str; pin: str

class Evento(BaseModel):
    doc: str; setor: str; proc: str
    paciente: Optional[str]=""; date: str; time: str
    obs: Optional[str]=""; ai: Optional[bool]=True
    pdf_filename: Optional[str]=""; pdf_data: Optional[str]=""
    duracao_min: Optional[int]=60

class EventoUpdate(BaseModel):
    doc: Optional[str]=None; setor: Optional[str]=None; proc: Optional[str]=None
    paciente: Optional[str]=None; date: Optional[str]=None; time: Optional[str]=None
    obs: Optional[str]=None; duracao_min: Optional[int]=None

class Medico(BaseModel):
    id: str; name: str; spec: Optional[str]=""; email: Optional[str]=""

class Setor(BaseModel):
    id: str; name: str
    color: Optional[str]="#CECBF6"; text_color: Optional[str]="#3C3489"
    endereco: Optional[str]=""
    tempo_manual: Optional[int]=0

class Memoria(BaseModel):
    id: str; texto: str
    icone: Optional[str]="ti-brain"; tipo: Optional[str]="aprendido"

class Usuario(BaseModel):
    id: str; nome: str; pin: str
    role: Optional[str]="medico"; email: Optional[str]=""
    medico_id: Optional[str]=""

class ChangePin(BaseModel):
    pin_atual: str; pin_novo: str

class Correcao(BaseModel):
    contexto: str  # texto original que o usuário digitou
    campo: str     # qual campo estava errado (ex: "setor", "horario", "medico")
    valor_errado: Optional[str] = ""
    valor_certo: str

# ── AUTH ROUTES ────────────────────────────────────────────
class SetupData(BaseModel):
    usuario_id: str; nome: str; pin: str
    gcal_calendar_id: Optional[str] = ""

class SignupData(BaseModel):
    usuario_id: str; nome: str; pin: str
    org_nome: Optional[str] = ""

@app.get("/api/me")
def me(user=Depends(auth)):
    return {k:v for k,v in user.items() if k!="pin_hash"}

# ── EVENTOS ────────────────────────────────────────────────
@app.patch("/api/eventos/{ev_id}/status")
async def update_status(ev_id: str, user=Depends(auth), status: str = "aguardando"):
    out = ana_data.sb_update_status(user, ev_id, status)
    if status == "cancelado":
        try:
            rows = sb_rest("GET", f"/appointments?id=eq.{ev_id}&select=created_by")
            criador = (rows[0].get("created_by") if rows else None) or user["id"]
            await gcal_delete(ev_id.replace("-", ""), criado_por_id=criador, org_id=user.get("org_id", "default"))
        except Exception as e:
            log.warning(f"GCal delete no cancelamento: {e}")
    return out

@app.get("/api/pacientes")
def list_pacientes(q: str = "", user=Depends(auth)):
    return ana_data.sb_list_pacientes(user, q)

@app.get("/api/eventos")
def list_eventos(user=Depends(auth)):
    return ana_data.sb_list_eventos(user)

@app.post("/api/eventos")
async def create_evento(ev: Evento, bg: BackgroundTasks, request: Request, user=Depends(auth)):
    out = ana_data.sb_create_evento(user, ev)
    org_id = user.get("org_id", "default")
    try:
        if get_user_google_token(user["id"]) or GCAL_CREDS:
            await gcal_create(ev.dict(), out.get("setor_nome") or "", criado_por_id=user["id"], org_id=org_id,
                             event_id=str(out.get("id", "")).replace("-", ""))
    except Exception as e:
        log.warning(f"GCal sync (modo supabase): {e}")
    if VAPID_PUBLIC_KEY:
        bg.add_task(push_all_org, org_id, "📅 Novo agendamento",
                    f"{ev.proc} — {ev.doc} · {fmtDate(ev.date)} {ev.time}")
    return out

@app.delete("/api/eventos/{ev_id}")
async def delete_evento(ev_id: str, bg: BackgroundTasks, user=Depends(auth)):
    try:
        rows = sb_rest("GET", f"/appointments?id=eq.{ev_id}&select=created_by")
        criador = (rows[0].get("created_by") if rows else None) or user["id"]
        await gcal_delete(ev_id.replace("-", ""), criado_por_id=criador, org_id=user.get("org_id", "default"))
    except Exception as e:
        log.warning(f"GCal delete na exclusão: {e}")
    return ana_data.sb_delete_evento(user, ev_id)

@app.put("/api/eventos/{ev_id}")
async def update_evento(ev_id: str, ev: EventoUpdate, bg: BackgroundTasks, user=Depends(auth)):
    return ana_data.sb_update_evento(user, ev_id, ev)

# ── MÉDICOS ────────────────────────────────────────────────
@app.get("/api/medicos")
def list_medicos(user=Depends(auth)):
    return ana_data.sb_list_medicos(user)

@app.post("/api/medicos")
def create_medico(m: Medico, user=Depends(auth)):
    return ana_data.sb_create_medico(user, m)

@app.delete("/api/medicos/{mid}")
def delete_medico(mid: str, user=Depends(auth)):
    return ana_data.sb_delete_medico(user, mid)

# ── SETORES ────────────────────────────────────────────────
@app.get("/api/setores")
def list_setores(user=Depends(auth)):
    return ana_data.sb_list_setores(user)

@app.post("/api/setores")
def create_setor(s: Setor, user=Depends(auth)):
    return ana_data.sb_create_setor(user, s)

@app.put("/api/setores/{sid}")
def update_setor(sid: str, s: Setor, user=Depends(auth)):
    return ana_data.sb_update_setor(user, sid, s)

@app.delete("/api/setores/{sid}")
def delete_setor(sid: str, user=Depends(auth)):
    return ana_data.sb_delete_setor(user, sid)

# ── MEMÓRIAS ───────────────────────────────────────────────
@app.get("/api/memorias")
def list_memorias(user=Depends(auth)):
    return ana_data.sb_list_memorias(user)

@app.post("/api/memorias")
def create_memoria(m: Memoria, user=Depends(auth)):
    return ana_data.sb_create_memoria(user, m)

@app.delete("/api/memorias/{mid}")
def delete_memoria(mid: str, user=Depends(auth)):
    return ana_data.sb_delete_memoria(user, mid)

@app.delete("/api/memorias")
def clear_memorias(user=Depends(auth)):
    return ana_data.sb_delete_memoria(user, "all")

# ── CORREÇÕES (aprendizado a partir de erros) ───────────────
@app.post("/api/correcoes")
def create_correcao(c_in: Correcao, user=Depends(auth)):
    return ana_data.sb_create_correcao(user, c_in)

@app.get("/api/correcoes")
def list_correcoes(user=Depends(auth)):
    return ana_data.sb_list_correcoes(user)

# ── CONTEXTO IA ────────────────────────────────────────────
@app.get("/api/contexto-ia")
def get_contexto(user=Depends(auth)):
    return ana_data.sb_contexto_ia(user)

# ── RELATÓRIOS ─────────────────────────────────────────────
@app.get("/api/relatorios/resumo")
def relatorio_resumo(user=Depends(auth)):
    return ana_data.sb_relatorio_resumo(user)

# ── MAPA CIRÚRGICO (organizado por sala/setor) ──────────────
@app.get("/api/mapa-cirurgico")
def mapa_cirurgico(data: str, user=Depends(auth)):
    return ana_data.sb_mapa_cirurgico(user, data)

# ── HISTÓRICO E LOGS ───────────────────────────────────────
@app.get("/api/historico")
def list_historico(user=Depends(auth)):
    return ana_data.sb_list_historico(user)

@app.get("/api/logs")
def list_logs(user=Depends(auth)):
    return ana_data.sb_list_logs(user)

# ── LEMBRETE DIÁRIO ──────────────────────────────────────────
def reminder_email_html(eventos_dia, data_str):
    rows = ""
    for e in eventos_dia:
        s = e.get("setor","")
        rows += f"""<tr>
          <td style="padding:5px 8px;border-bottom:0.5px solid #eee">{e['time']}</td>
          <td style="padding:5px 8px;border-bottom:0.5px solid #eee">{s}</td>
          <td style="padding:5px 8px;border-bottom:0.5px solid #eee">{e['proc']}</td>
          <td style="padding:5px 8px;border-bottom:0.5px solid #eee">{e.get('paciente','—')}</td>
          <td style="padding:5px 8px;border-bottom:0.5px solid #eee;font-size:11px;color:#888">{e['doc']}</td>
        </tr>"""
    return f"""<div style="font-family:Arial;max-width:560px;margin:0 auto;padding:20px">
      <div style="background:#6C63D4;color:#fff;border-radius:10px 10px 0 0;padding:14px 18px">
        <b>A.N.A · Resumo de {data_str}</b></div>
      <div style="background:#f9f9ff;border:1px solid #E4E4EF;border-radius:0 0 10px 10px;padding:18px">
        <p style="font-size:13px;color:#555;margin-bottom:10px">{len(eventos_dia)} procedimento(s) agendado(s) para o dia.</p>
        <table style="width:100%;font-size:12px;border-collapse:collapse">
          <tr><th style="text-align:left;padding:5px 8px;background:#EEEDFE;color:#3C3489">Hora</th>
              <th style="text-align:left;padding:5px 8px;background:#EEEDFE;color:#3C3489">Setor</th>
              <th style="text-align:left;padding:5px 8px;background:#EEEDFE;color:#3C3489">Procedimento</th>
              <th style="text-align:left;padding:5px 8px;background:#EEEDFE;color:#3C3489">Paciente</th>
              <th style="text-align:left;padding:5px 8px;background:#EEEDFE;color:#3C3489">Médico</th></tr>
          {rows}
        </table>
        <p style="font-size:10px;color:#aaa;margin-top:14px">A.N.A · Secretária Virtual — lembrete automático</p>
      </div></div>"""

async def run_daily_reminder(org_id: Optional[str] = None):
    """Resumo da agenda de amanhã por grupo: push para os dispositivos registrados
    e email para médicos com email (quando SMTP configurado)."""
    amanha = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
    amanha_str = (datetime.now() + timedelta(days=1)).strftime("%d/%m/%Y")
    if org_id:
        grupos = [{"id": org_id}]
    else:
        # só grupos com alguém inscrito em push (evita varrer grupos inativos)
        subs = sb_rest("GET", "/ana_push_subscriptions?select=group_id")
        grupos = [{"id": g} for g in {s["group_id"] for s in subs}]
    total_push, total_email = 0, 0
    for g in grupos:
        gid = g["id"]
        appts = sb_rest("GET", f"/appointments?group_id=eq.{gid}&appointment_date=eq.{amanha}"
                               f"&status=neq.cancelled&select=*&order=appointment_time")
        if not appts:
            continue
        n = len(appts)
        await push_all_org(gid, f"🩺 A.N.A · Agenda de {amanha_str}",
                           f"{n} procedimento{'s' if n != 1 else ''} agendado{'s' if n != 1 else ''} para amanhã", "/")
        total_push += 1
        # email por médico (opcional, requer SMTP e doctors.phone/email não existe — via profiles dos vinculados)
        if SMTP_HOST:
            docs = sb_rest("GET", f"/doctors?group_id=eq.{gid}&user_id=not.is.null&select=id,name,user_id")
            if docs:
                uids = ",".join(f'"{d["user_id"]}"' for d in docs)
                emails = {p["id"]: p.get("email") for p in sb_rest("GET", f"/profiles?id=in.({uids})&select=id,email")}
                secs = {s["id"]: s["name"] for s in sb_rest("GET", f"/sectors?group_id=eq.{gid}&select=id,name")}
                for d in docs:
                    meus = [a for a in appts if a.get("doctor_id") == d["id"]]
                    email = emails.get(d["user_id"])
                    if not meus or not email:
                        continue
                    evs = [{"time": str(a.get("appointment_time") or "")[:5], "proc": a.get("procedure") or "",
                            "paciente": a.get("patient_name") or "", "setor": secs.get(a.get("sector_id") or "", ""),
                            "doc": d["name"], "obs": a.get("notes") or ""} for a in meus]
                    try:
                        await send_email(email, f"A.N.A · Sua agenda de {amanha_str}",
                                         reminder_email_html(evs, amanha_str))
                        total_email += 1
                    except Exception as e:
                        log.warning(f"email lembrete: {e}")
    return {"sent": total_email, "push_groups": total_push}

@app.post("/api/lembrete-diario")
async def trigger_daily_reminder(user=Depends(auth)):
    """Dispara manualmente o envio do lembrete do dia seguinte (admin) — só para o grupo do usuário."""
    if user["role"] != "admin":
        raise HTTPException(403, "Acesso negado.")
    result = await run_daily_reminder(org_id=user.get("org_id","default"))
    return result

@app.on_event("startup")
async def start_scheduler():
    import asyncio
    async def scheduler_loop():
        last_run_date = None
        while True:
            now = datetime.now()
            # Dispara uma vez por dia, próximo das 18h
            if now.hour == 18 and last_run_date != now.date():
                try:
                    await run_daily_reminder()
                except Exception as e:
                    log.error(f"Erro no lembrete diário: {e}")
                last_run_date = now.date()
            await asyncio.sleep(60 * 10)  # checa a cada 10 minutos
    asyncio.create_task(scheduler_loop())

# ── CHAT PROXY (Groq — gratuito, evita CORS) ──────────────
def _extract_pdf_gemini(pdf_b64: str) -> str:
    """Usa o Gemini Flash para extrair e estruturar o conteúdo do PDF nativamente.
    Muito superior ao pypdf para PDFs com tabelas, colunas e formatação complexa."""
    try:
        prompt = """Você é um assistente de extração de dados médicos. Analise este PDF e extraia TODAS as informações relevantes de forma estruturada.

Retorne um texto claro com:
- Data da agenda
- Para cada paciente: nome completo, data/hora do procedimento, procedimento exato, se é COM ANESTESIA ou SEM ANESTESIA, convênio, observações relevantes
- Destaque claramente quais procedimentos requerem anestesia

Seja preciso e inclua todos os pacientes listados. Não omita nenhum dado."""

        payload = json.dumps({
            "contents": [{
                "parts": [
                    {
                        "inline_data": {
                            "mime_type": "application/pdf",
                            "data": pdf_b64
                        }
                    },
                    {"text": prompt}
                ]
            }],
            "generationConfig": {
                "temperature": 0.1,
                "maxOutputTokens": 8000,
                "thinkingConfig": {"thinkingBudget": 0}
            }
        }).encode("utf-8")

        data = _gemini_request(payload, timeout=30)

        text = data["candidates"][0]["content"]["parts"][0]["text"]
        log.info(f"Gemini PDF extraction: {len(text)} chars extraídos")
        return text
    except Exception as e:
        log.error(f"Gemini PDF extraction erro: {e}")
        return ""

def _extract_image_ocrspace(image_b64: str, mime_type: str = "image/png") -> str:
    """Fallback: OCR.space para extração de texto de imagens. Gratuito até 25k req/mês."""
    if not OCR_SPACE_API_KEY:
        return ""
    try:
        data_url = f"data:{mime_type};base64,{image_b64}"
        body = urllib.parse.urlencode({
            "base64Image": data_url,
            "apikey": OCR_SPACE_API_KEY,
            "language": "por",
            "isOverlayRequired": "false",
            "detectOrientation": "true",
            "scale": "true",
            "OCREngine": "2",
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://api.ocr.space/parse/image",
            data=body,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.loads(r.read())
        if data.get("IsErroredOnProcessing"):
            log.error(f"OCR.space erro: {data.get('ErrorMessage')}")
            return ""
        results = data.get("ParsedResults", [])
        if not results:
            return ""
        text = "\n".join(r.get("ParsedText", "") for r in results).strip()
        log.info(f"OCR.space extraído: {len(text)} chars")
        return text
    except Exception as e:
        log.error(f"OCR.space erro: {e}")
        return ""

def _gemini_request(payload_bytes: bytes, timeout: int = 60):
    """POST no Gemini nativo compatível com chaves antigas (AIza) e novas (AQ. Auth keys).
    Tenta o header x-goog-api-key; se a chave nova for recusada (401/403), tenta Authorization: Bearer."""
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"
    tentativas = [{"x-goog-api-key": GEMINI_API_KEY}]
    if GEMINI_API_KEY.startswith("AQ."):
        tentativas.append({"Authorization": f"Bearer {GEMINI_API_KEY}"})
    ultimo = None
    for auth_h in tentativas:
        headers = {"Content-Type": "application/json", "User-Agent": "ANA-Secretaria/1.0", **auth_h}
        req = urllib.request.Request(url, data=payload_bytes, method="POST", headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return json.loads(resp.read())
        except urllib.error.HTTPError as e:
            if e.code in (401, 403) and len(tentativas) > 1 and auth_h is not tentativas[-1]:
                ultimo = e
                log.warning(f"Gemini {e.code} com {list(auth_h)[0]} — tentando Authorization: Bearer")
                continue
            raise
    raise ultimo


def _extract_image_gemini(image_b64: str, mime_type: str = "image/jpeg") -> str:
    """Usa o Gemini Flash para extrair informações de agenda de uma imagem."""
    if not GEMINI_API_KEY:
        return ""
    try:
        prompt = """Você é um assistente de extração de dados médicos. Analise esta imagem de agenda médica e extraia TODAS as informações relevantes de forma estruturada.

Retorne um texto claro com:
- Data da agenda (se visível)
- Para cada paciente: nome completo, data/hora do procedimento, procedimento exato, se é COM ANESTESIA ou SEM ANESTESIA, convênio, observações relevantes
- Destaque claramente quais procedimentos requerem anestesia

Seja preciso e inclua todos os pacientes listados. Não omita nenhum dado."""

        payload = json.dumps({
            "contents": [{
                "parts": [
                    {"inline_data": {"mime_type": mime_type, "data": image_b64}},
                    {"text": prompt}
                ]
            }],
            "generationConfig": {"temperature": 0.1, "maxOutputTokens": 8000,
                                 "thinkingConfig": {"thinkingBudget": 0}}
        }).encode("utf-8")

        data = _gemini_request(payload, timeout=45)

        # Verifica se há candidatos válidos
        candidates = data.get("candidates", [])
        if not candidates:
            log.error(f"Gemini image: sem candidatos. Resposta: {json.dumps(data)[:300]}")
            return ""
        parts = candidates[0].get("content", {}).get("parts", [])
        if not parts:
            log.error(f"Gemini image: sem parts. Candidato: {json.dumps(candidates[0])[:300]}")
            return ""
        text = parts[0].get("text", "")
        fr = candidates[0].get("finishReason", "")
        if fr == "MAX_TOKENS":
            log.warning("Gemini image: extração TRUNCADA por limite de tokens — lista pode estar incompleta")
        log.info(f"Gemini image: {len(text)} chars extraídos, {text.count(chr(10))+1} linhas, finish={fr}")
        if not text:
            log.error(f"Gemini image: texto vazio. Parts: {json.dumps(parts)[:300]}")
            return ""
        log.info(f"Gemini image extraction: {len(text)} chars extraídos, mime={mime_type}")
        return text
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        log.error(f"Gemini image HTTP {e.code}: {body[:400]}")
        if e.code == 429:
            raise  # Re-lança 429 para o caller tentar o fallback
        return ""
    except Exception as e:
        log.error(f"Gemini image extraction erro: {type(e).__name__}: {e}")
        return ""


def _extract_pdf_text(pdf_b64: str, max_chars: int = 20000) -> str:
    """Extrai texto de um PDF em base64 via pypdf (fallback quando sem Gemini API key)."""
    if not PDF_SUPPORT or not pdf_b64:
        return ""
    try:
        pdf_bytes = base64.b64decode(pdf_b64)
        reader = PdfReader(io.BytesIO(pdf_bytes))
        text = ""
        for page in reader.pages:
            # modo layout preserva as colunas de tabelas (mapas cirúrgicos) — essencial para a IA
            try:
                t = page.extract_text(extraction_mode="layout") or ""
            except Exception:
                t = ""
            if not t.strip():
                t = page.extract_text() or ""
            text += t
            if len(text) >= max_chars:
                break
        result = text[:max_chars].strip()
        log.info(f"pypdf extração: {len(result)} chars de {len(reader.pages)} página(s)")
        return result
    except Exception as e:
        log.error(f"Erro ao extrair texto do PDF: {e}")
        return ""

def extract_pdf(pdf_b64: str) -> str:
    """Extrai conteúdo do PDF: usa Gemini se disponível, senão pypdf."""
    if GEMINI_API_KEY and pdf_b64:
        result = _extract_pdf_gemini(pdf_b64)
        if result:
            return result
        log.warning("Gemini falhou na extração do PDF, tentando pypdf...")
    return _extract_pdf_text(pdf_b64)

class ChatRequest(BaseModel):
    system: str
    messages: List[Any]
    max_tokens: Optional[int] = None

@app.post("/api/chat")
def chat_proxy(req: ChatRequest, user=Depends(auth)):
    """Cascata de motores: Gemini 2.5 Flash → Groq → Cerebras (os configurados, nessa ordem)."""
    motores = []
    if GEMINI_API_KEY: motores.append("gemini")
    if GROQ_KEY: motores.append("groq")
    if CEREBRAS_API_KEY: motores.append("cerebras")
    if not motores:
        raise HTTPException(500, "Nenhum motor de IA configurado (GEMINI_API_KEY / GROQ_API_KEY / CEREBRAS_API_KEY).")
    log.info(f"Chat proxy [cascata: {'→'.join(motores)}]: system={len(req.system)} chars, msgs={len(req.messages)}")

    # Converte formato Anthropic (content pode ser string ou lista de partes) → OpenAI/Groq
    tem_anexo = False
    openai_messages = [{"role": "system", "content": req.system}]
    for msg in req.messages:
        content = msg.get("content", "")
        role = msg.get("role", "user")
        if isinstance(content, str):
            text = content
        elif isinstance(content, list):
            text_parts = []
            for part in content:
                if part.get("type") == "text":
                    text_parts.append(part.get("text", ""))
                elif part.get("type") == "document":
                    tem_anexo = True
                    src = part.get("source", {})
                    pdf_b64 = src.get("data", "")
                    extracted = extract_pdf(pdf_b64)
                    if extracted:
                        text_parts.append(f"[Conteúdo extraído do PDF anexado]\n{extracted}")
                    else:
                        text_parts.append("[Aviso: não foi possível ler o conteúdo do PDF anexado. "
                                          "Peça ao usuário para descrever o pedido médico em texto.]")
                elif part.get("type") == "image":
                    tem_anexo = True
                    src = part.get("source", {})
                    img_b64 = src.get("data", "")
                    mime = src.get("media_type", "image/jpeg")
                    extracted = ""
                    # Tenta Gemini primeiro, cai para OCR.space se falhar
                    if img_b64 and GEMINI_API_KEY:
                        try:
                            extracted = _extract_image_gemini(img_b64, mime)
                        except urllib.error.HTTPError as e:
                            if e.code == 429:
                                log.warning("Gemini quota esgotada — tentando OCR.space como fallback")
                            else:
                                log.error(f"Gemini image erro {e.code}")
                    if not extracted and img_b64 and OCR_SPACE_API_KEY:
                        extracted = _extract_image_ocrspace(img_b64, mime)
                        if extracted:
                            log.info("OCR.space usado como fallback para imagem")
                    if extracted:
                        text_parts.append(f"[Conteúdo extraído da imagem anexada]\n{extracted}")
                    else:
                        text_parts.append("[Aviso: não foi possível extrair informações da imagem. Configure GEMINI_API_KEY ou OCR_SPACE_API_KEY no Railway.]")
            text = "\n".join(text_parts)
        else:
            text = str(content)
        openai_messages.append({"role": role, "content": text})

    # max_tokens dinâmico: reserva grande só quando há anexo (AGENDAR_MULTIPLOS de listas)
    limite = 6000 if tem_anexo else 2000
    max_tk = min(req.max_tokens or limite, limite)

    limite = 6000 if tem_anexo else 2000
    max_tk = min(req.max_tokens or limite, limite)

    def _openai_compat(nome, url, key, model):
        payload = json.dumps({"model": model, "max_tokens": max_tk,
                              "messages": openai_messages, "temperature": 0.3}).encode("utf-8")
        http_req = urllib.request.Request(url, data=payload, method="POST",
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {key}",
                     "User-Agent": "ANA-Secretaria/1.0"})
        with urllib.request.urlopen(http_req, timeout=60) as resp:
            j = json.loads(resp.read())
        return j["choices"][0]["message"]["content"]

    def _gemini_chat():
        contents = []
        for m in openai_messages:
            if m["role"] == "system":
                continue
            contents.append({"role": "model" if m["role"] == "assistant" else "user",
                             "parts": [{"text": m["content"]}]})
        payload = json.dumps({
            "systemInstruction": {"parts": [{"text": req.system}]},
            "contents": contents,
            "generationConfig": {"temperature": 0.3,
                                 # com anexo: libera raciocínio (2048) p/ organizar escalas; o teto acomoda o thinking
                                 "maxOutputTokens": max_tk + (2048 if tem_anexo else 0),
                                 "responseMimeType": "application/json",
                                 "thinkingConfig": {"thinkingBudget": 2048 if tem_anexo else 0}},
        }).encode("utf-8")
        j = _gemini_request(payload, timeout=60)
        cand = (j.get("candidates") or [{}])[0]
        parts = (cand.get("content") or {}).get("parts") or []
        text = "".join(p.get("text", "") for p in parts)
        if not text:
            raise RuntimeError(f"Gemini sem texto (finish={cand.get('finishReason')})")
        return text

    erros_cascata = []
    ultimo_erro = None
    for idx, motor in enumerate(motores):
        ultimo = idx == len(motores) - 1
        tentativas = 2 if ultimo else 1   # no último motor, insiste uma vez a mais no 429
        for t in range(tentativas):
            try:
                if motor == "gemini":
                    text = _gemini_chat()
                elif motor == "groq":
                    text = _openai_compat("groq", "https://api.groq.com/openai/v1/chat/completions",
                                          GROQ_KEY, "llama-3.3-70b-versatile")
                else:
                    text = _openai_compat("cerebras", "https://api.cerebras.ai/v1/chat/completions",
                                          CEREBRAS_API_KEY, CEREBRAS_MODEL)
                log.info(f"Chat proxy: resposta OK ({motor})")
                return {"content": [{"type": "text", "text": text}]}
            except urllib.error.HTTPError as e:
                body = e.read().decode("utf-8", errors="ignore")
                ultimo_erro = f"{motor} {e.code}: {body[:200]}"
                erros_cascata.append(f"{motor}: HTTP {e.code}")
                if e.code == 429 and ultimo and t == 0:
                    m = re.search(r"try again in ([\d.]+)s", body)
                    espera = min(float(m.group(1)) + 0.4, 8.0) if m else 3.0
                    log.warning(f"{motor} 429 (último motor) — aguardando {espera:.1f}s")
                    time.sleep(espera)
                    continue
                log.warning(f"Motor {motor} falhou ({e.code}) — {'tentando o próximo' if not ultimo else 'sem mais motores'}")
                break
            except Exception as e:
                ultimo_erro = f"{motor}: {e}"
                erros_cascata.append(f"{motor}: {str(e)[:80]}")
                log.warning(f"Motor {motor} falhou ({e}) — {'tentando o próximo' if not ultimo else 'sem mais motores'}")
                break
    log.error(f"Todos os motores falharam: {' | '.join(erros_cascata)} — último detalhe: {ultimo_erro}")
    if ultimo_erro and "429" in ultimo_erro:
        raise HTTPException(502, "As IAs estão no limite de uso por minuto — aguarde alguns segundos e tente de novo.")
    raise HTTPException(502, f"Falha nos motores de IA — {'; '.join(erros_cascata)}. Confira as chaves no Railway (logs têm o detalhe).")
@app.get("/api/push/vapid-key")
def get_vapid_key():
    """Retorna a chave pública VAPID para o frontend registrar subscriptions."""
    if not VAPID_PUBLIC_KEY:
        raise HTTPException(400, "Web Push não configurado no servidor.")
    return {"public_key": VAPID_PUBLIC_KEY}

async def push_all_org(org_id: str, title: str, body: str, url: str = "/"):
    """Envia push para todos os dispositivos inscritos do grupo (limpa inscrições mortas)."""
    if not VAPID_PRIVATE_KEY:
        return
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        log.warning("pywebpush não instalado — push ignorado")
        return
    subs = sb_rest("GET", f"/ana_push_subscriptions?group_id=eq.{org_id}&select=id,subscription")
    payload = json.dumps({"title": title, "body": body, "url": url})
    enviados = 0
    for s in subs:
        try:
            webpush(subscription_info=s["subscription"], data=payload,
                    vapid_private_key=VAPID_PRIVATE_KEY,
                    vapid_claims={"sub": f"mailto:{VAPID_CLAIMS_EMAIL}"})
            enviados += 1
        except WebPushException as e:
            code = getattr(getattr(e, "response", None), "status_code", 0)
            if code in (404, 410):  # inscrição morta — remove
                sb_rest("DELETE", f"/ana_push_subscriptions?id=eq.{s['id']}")
            else:
                log.warning(f"push falhou: {e}")
        except Exception as e:
            log.warning(f"push erro: {e}")
    if subs:
        log.info(f"Push '{title}' → {enviados}/{len(subs)} dispositivos")


@app.post("/api/push/subscribe")
async def push_subscribe(request: Request, user=Depends(auth)):
    """Salva a subscription de push do dispositivo do usuário."""
    body = await request.json()
    sub_json = json.dumps(body)
    sub_id = secrets.token_hex(16)
    org_id = user.get("org_id", "default")
    # remove inscrições anteriores do usuário e insere a nova
    sb_rest("DELETE", f"/ana_push_subscriptions?user_id=eq.{user['id']}")
    sb_rest("POST", "/ana_push_subscriptions",
            {"user_id": user["id"], "group_id": org_id, "subscription": body})
    return {"ok": True}

@app.delete("/api/push/subscribe")
def push_unsubscribe(user=Depends(auth)):
    """Remove todas as subscriptions do usuário."""
    sb_rest("DELETE", f"/ana_push_subscriptions?user_id=eq.{user['id']}")
    return {"ok": True}

@app.get("/api/push/status")
def push_status(user=Depends(auth)):
    rows = sb_rest("GET", f"/ana_push_subscriptions?user_id=eq.{user['id']}&select=id")
    return {"enabled": len(rows) > 0, "vapid_configured": bool(VAPID_PUBLIC_KEY)}

@app.post("/api/push/test")
async def push_test(user=Depends(auth)):
    """Envia uma notificação de teste para o usuário atual."""
    rows = sb_rest("GET", f"/ana_push_subscriptions?user_id=eq.{user['id']}&select=subscription")
    subs = [{"subscription": json.dumps(r["subscription"])} for r in rows]
    if not subs:
        raise HTTPException(400, "Nenhum dispositivo registrado para este usuário.")
    sent = 0
    for sub in subs:
        info = json.loads(sub["subscription"])
        if send_push(info, "🩺 A.N.A · Teste", "Notificações funcionando!", "/"):
            sent += 1
    return {"ok": True, "sent": sent}

# ── CONFIGURAÇÃO DO GOOGLE CALENDAR (via app) ───────────────
class GCalConfig(BaseModel):
    calendar_id: str

@app.get("/api/config/gcal")
def get_gcal_config(user=Depends(auth)):
    org_id = user.get("org_id","default")
    return {"calendar_id": get_gcal_id(org_id), "configurado_via": "app" if get_config("gcal_calendar_id", org_id=org_id) else "env_var_ou_padrao"}

@app.post("/api/config/gcal")
def set_gcal_config(cfg: GCalConfig, user=Depends(auth)):
    if user["role"] != "admin":
        raise HTTPException(403, "Apenas administradores podem alterar essa configuração.")
    org_id = user.get("org_id","default")
    set_config("gcal_calendar_id", cfg.calendar_id.strip(), org_id=org_id)
    db_log("INFO", f"Google Calendar ID alterado para: {cfg.calendar_id}", usuario=user["id"], org_id=org_id)
    return {"ok": True, "calendar_id": cfg.calendar_id.strip()}

# ── DADOS DO GRUPO (nome institucional para cabeçalhos de relatórios) ──
class OrgInfo(BaseModel):
    nome: str

@app.get("/api/org/info")
def get_org_info(user=Depends(auth)):
    return ana_data.sb_get_org_info(user)

@app.post("/api/org/info")
def set_org_info(info: OrgInfo, user=Depends(auth)):
    return ana_data.sb_set_org_info(user, info)

@app.get("/api/config/gcal/list")
def list_gcal_calendars(user=Depends(auth)):
    """Lista os calendários acessíveis pela conta de serviço configurada — ajuda a escolher o ID certo."""
    if not GCAL_CREDS:
        raise HTTPException(400, "Google Calendar não está configurado (GCAL_CREDENTIALS ausente).")
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        creds = Credentials.from_service_account_info(
            json.loads(GCAL_CREDS),
            scopes=["https://www.googleapis.com/auth/calendar.readonly"])
        svc = build("calendar", "v3", credentials=creds)
        result = svc.calendarList().list().execute()
        cals = [{"id": c.get("id"), "summary": c.get("summary"), "primary": c.get("primary", False)}
                for c in result.get("items", [])]
        return {"calendars": cals}
    except Exception as e:
        log.error(f"Erro ao listar calendários: {e}")
        raise HTTPException(502, f"Não foi possível listar calendários: {str(e)[:200]}")

# ── OAUTH GOOGLE (calendário pessoal por usuário) ───────────
@app.get("/api/oauth/google/status")
def oauth_google_status(user=Depends(auth)):
    row = _gtok_get(user["id"])
    connected = bool(row.get("refresh_token"))
    return {"connected": connected, "email": row.get("email",""),
            "oauth_disponivel": bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET and APP_BASE_URL)}

@app.get("/api/oauth/google/start")
def oauth_google_start(request: Request, user=Depends(auth)):
    if not (GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET and APP_BASE_URL):
        raise HTTPException(400, "OAuth do Google não está configurado no servidor. Configure GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET e APP_BASE_URL no Railway.")
    # state = token de sessão atual, para sabermos a qual usuário vincular no callback
    token = request.headers.get("X-Token","") or request.query_params.get("token","")
    if not token:
        authz = request.headers.get("Authorization", "")
        if authz.startswith("Bearer "):
            token = authz[7:]
    if not token:
        raise HTTPException(401, "Token ausente.")
    url = build_oauth_url(state=token)
    return {"auth_url": url}

@app.get("/api/oauth/google/callback")
def oauth_google_callback(code: str = "", state: str = "", error: str = ""):
    if error:
        return HTMLResponse(f"<html><body style='font-family:sans-serif;text-align:center;padding:40px'><h3>Autorização cancelada</h3><p>{error}</p><script>setTimeout(()=>window.close(),2000)</script></body></html>")
    # state = JWT do Supabase do usuário que iniciou a conexão
    user = None
    claims = validate_supabase_jwt(state)
    if claims and claims.get("sub"):
        meta = claims.get("user_metadata") or {}
        nome = meta.get("full_name") or meta.get("name") or (claims.get("email", "").split("@")[0] or claims["sub"][:8])
        user = {"id": claims["sub"], "nome": nome}
    if not user:
        return HTMLResponse("<html><body style='font-family:sans-serif;text-align:center;padding:40px'><h3>Sessão inválida ou expirada</h3><p>Volte ao app e tente novamente.</p></body></html>")
    try:
        tok = exchange_oauth_code(code)
        access_token = tok.get("access_token","")
        refresh_token = tok.get("refresh_token","")
        expires_in = tok.get("expires_in", 3600)
        expiry = (datetime.now() + timedelta(seconds=expires_in - 60)).isoformat()

        # Busca o email da conta Google conectada
        email = ""
        try:
            req = urllib.request.Request("https://www.googleapis.com/oauth2/v2/userinfo",
                                          headers={"Authorization": f"Bearer {access_token}"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                info = json.loads(resp.read())
                email = info.get("email","")
        except Exception:
            pass

        if refresh_token:
            _gtok_save(user["id"], access_token=access_token, refresh_token=refresh_token, expiry=expiry, email=email)
        else:
            # Google só manda refresh_token na primeira autorização; se já existir, mantém o antigo
            _gtok_save(user["id"], access_token=access_token, expiry=expiry, email=email)
        db_log("INFO", f"Google Calendar pessoal conectado: {email}", usuario=user["id"])
        return HTMLResponse(f"""<html><body style='font-family:sans-serif;text-align:center;padding:40px'>
            <h3>✅ Calendário conectado!</h3><p>{email}</p>
            <p style='color:#888;font-size:13px'>Pode fechar esta janela.</p>
            <script>setTimeout(()=>{{window.close();if(window.opener)window.opener.location.reload();}},1500)</script>
            </body></html>""")
    except Exception as e:
        log.error(f"Erro no callback OAuth: {e}")
        return HTMLResponse(f"<html><body style='font-family:sans-serif;text-align:center;padding:40px'><h3>Erro ao conectar</h3><p>{str(e)[:200]}</p></body></html>")

@app.post("/api/oauth/google/disconnect")
def oauth_google_disconnect(user=Depends(auth)):
    _gtok_clear(user["id"])
    db_log("INFO", "Google Calendar pessoal desconectado", usuario=user["id"])
    return {"ok": True}

# ── HEALTH ─────────────────────────────────────────────────
@app.get("/api/public-config")
def public_config():
    """Config pública para o frontend inicializar o supabase-js (anon key é pública por design)."""
    return {"supabase_url": SUPABASE_URL, "supabase_anon_key": SUPABASE_ANON_KEY,
            "supabase_auth": bool(SUPABASE_URL and SUPABASE_ANON_KEY),
            "medss_url": os.environ.get("MEDSS_URL", "")}

@app.get("/api/escala")
def get_escala(mes: str = "", user=Depends(auth)):
    """Escala mensal (shifts do MedSS) — disponível apenas no backend Supabase."""
    if not mes:
        mes = datetime.now().strftime("%Y-%m")
    return ana_data.sb_escala(user, mes)

class Plantao(BaseModel):
    medico: str
    data: str
    turno: str          # morning | afternoon | night
    meio: bool = False
    hnt: bool = False
    setor: str = ""
    creditos_especiais: list = []
    repetir_ate: str = ""

@app.post("/api/escala/plantao")
def criar_plantao(p: Plantao, user=Depends(auth)):
    return ana_data.sb_create_plantao(user, p)

@app.delete("/api/escala/plantao/{shift_id}")
def remover_plantao(shift_id: str, user=Depends(auth)):
    return ana_data.sb_delete_plantao(user, shift_id)

class SwapRequest(BaseModel):
    meu_shift_id: str
    alvo_shift_id: str
    mensagem: str = ""

class SwapResposta(BaseModel):
    aceitar: bool

@app.get("/api/creditos")
def get_creditos(mes: str = "", user=Depends(auth)):
    if not mes: mes = datetime.now().strftime("%Y-%m")
    return ana_data.sb_creditos(user, mes)

@app.get("/api/creditos/config")
def get_credit_config(user=Depends(auth)):
    return ana_data.sb_credit_settings(user)

@app.post("/api/creditos/config")
async def save_credit_config(request: Request, user=Depends(auth)):
    body = await request.json()
    return ana_data.sb_credit_settings_save(user, body)

@app.get("/api/creditos/export")
def export_creditos(mes: str = "", user=Depends(auth)):
    """Extrato de créditos do mês em Excel (.xlsx)."""
    if not mes:
        mes = datetime.now().strftime("%Y-%m")
    dados = ana_data.sb_creditos(user, mes)
    org = ana_data.sb_get_org_info(user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()

    FONT = "Arial"
    thin = Border(bottom=Side(style="thin", color="D9D9D9"))
    head_fill = PatternFill("solid", fgColor="4338CA")
    head_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)

    # ── Aba 1: Resumo por médico ──
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = f"Extrato de Créditos — {org['nome']}"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ano, m = mes.split("-")
    MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho",
             "Agosto","Setembro","Outubro","Novembro","Dezembro"]
    ws["A2"] = f"{MESES[int(m)-1]} de {ano}"
    ws["A2"].font = Font(name=FONT, size=11, color="666666")
    headers = ["Médico", "Plantões", "Detalhe", "Créditos"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = head_font; c.fill = head_fill
        c.alignment = Alignment(horizontal="center" if col > 1 else "left")
    r = 5
    total_geral = 0.0
    for med in dados["medicos"]:
        ws.cell(row=r, column=1, value=med["medico"]).font = Font(name=FONT)
        ws.cell(row=r, column=2, value=med["plantoes"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).font = Font(name=FONT)
        det = " · ".join(f"{v}× {k}" for k, v in med["detalhe"].items())
        ws.cell(row=r, column=3, value=det).font = Font(name=FONT, size=9, color="666666")
        c = ws.cell(row=r, column=4, value=med["total"])
        c.font = Font(name=FONT, bold=True); c.number_format = "0.0"
        c.alignment = Alignment(horizontal="center")
        for col in range(1, 5): ws.cell(row=r, column=col).border = thin
        total_geral += med["total"]; r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(name=FONT, bold=True)
    c = ws.cell(row=r, column=4, value=round(total_geral, 2))
    c.font = Font(name=FONT, bold=True); c.number_format = "0.0"
    c.alignment = Alignment(horizontal="center")
    for col, w in enumerate([28, 10, 48, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Aba 2: Valores configurados ──
    ws2 = wb.create_sheet("Valores")
    cs = dados["labels"]
    ws2["A1"] = "Valores de crédito configurados"
    ws2["A1"].font = Font(name=FONT, bold=True, size=12)
    for col, h in enumerate(["Tipo", "Créditos"], 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = head_font; c.fill = head_fill
    r = 4
    for k in ["morning", "afternoon", "night", "saturday", "sunday", "hnt_ambulatory"]:
        ws2.cell(row=r, column=1, value=cs.get(k + "_label", k)).font = Font(name=FONT)
        c = ws2.cell(row=r, column=2, value=float(cs.get(k + "_credit", 0)))
        c.font = Font(name=FONT); c.number_format = "0.0"
        r += 1
    ws2.column_dimensions["A"].width = 24; ws2.column_dimensions["B"].width = 10

    import io
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from fastapi.responses import StreamingResponse
    fname = f"creditos-{mes}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})

@app.get("/api/creditos/tipos")
def list_credit_types(user=Depends(auth)):
    return ana_data.sb_custom_types_list(user)

@app.post("/api/creditos/tipos")
async def create_credit_type(request: Request, user=Depends(auth)):
    body = await request.json()
    return ana_data.sb_custom_types_create(user, body)

@app.delete("/api/creditos/tipos/{type_id}")
def delete_credit_type(type_id: str, user=Depends(auth)):
    return ana_data.sb_custom_types_delete(user, type_id)

@app.post("/api/escala/troca")
def criar_troca(s: SwapRequest, bg: BackgroundTasks, user=Depends(auth)):
    out = ana_data.sb_swap_create(user, s)
    if VAPID_PUBLIC_KEY:
        bg.add_task(push_all_org, user.get("org_id", "default"),
                    "🔄 Proposta de troca de plantão",
                    f"{user['nome']} propôs uma troca — veja na aba Escala")
    return out

class Anuncio(BaseModel):
    shift_id: str
    mensagem: str = ""

@app.post("/api/escala/anunciar")
def anunciar_plantao(a: Anuncio, bg: BackgroundTasks, user=Depends(auth)):
    out = ana_data.sb_swap_announce(user, a.shift_id, a.mensagem)
    if VAPID_PUBLIC_KEY:
        TUR = {"morning": "manhã", "afternoon": "tarde", "night": "noite"}
        bg.add_task(push_all_org, user.get("org_id", "default"),
                    "📢 Plantão disponível",
                    f"{user['nome']} anunciou o plantão de {fmtDate(out['data'])} ({TUR.get(out['turno'], out['turno'])}) — quem puder assumir, veja na Escala")
    return out

@app.post("/api/escala/troca/{swap_id}/assumir")
def assumir_plantao(swap_id: str, bg: BackgroundTasks, user=Depends(auth)):
    out = ana_data.sb_swap_assume(user, swap_id)
    if VAPID_PUBLIC_KEY:
        bg.add_task(push_all_org, user.get("org_id", "default"),
                    "✅ Plantão assumido",
                    f"{user['nome']} assumiu o plantão anunciado — escala atualizada")
    return out

@app.get("/api/escala/trocas")
def listar_trocas(user=Depends(auth)):
    return ana_data.sb_swap_list(user)

@app.post("/api/escala/troca/{swap_id}/responder")
def responder_troca(swap_id: str, r: SwapResposta, bg: BackgroundTasks, user=Depends(auth)):
    out = ana_data.sb_swap_respond(user, swap_id, r.aceitar)
    if VAPID_PUBLIC_KEY:
        bg.add_task(push_all_org, user.get("org_id", "default"),
                    "🔄 Troca de plantão " + ("aceita ✅" if r.aceitar else "recusada"),
                    f"Resposta de {user['nome']} — escala atualizada" if r.aceitar else f"{user['nome']} recusou a proposta")
    return out

@app.delete("/api/escala/troca/{swap_id}")
def cancelar_troca(swap_id: str, user=Depends(auth)):
    return ana_data.sb_swap_cancel(user, swap_id)

class CodigoAcesso(BaseModel):
    codigo: str

class SolicitacaoResposta(BaseModel):
    aprovar: bool

@app.get("/api/grupos/meus")
def meus_grupos(request: Request):
    """Grupos do usuário com nomes — para o hub pós-login."""
    u = auth_jwt_only(request)
    memberships = get_memberships(u["id"])
    # solicitações pendentes que eu fiz (importante também para quem ainda não tem grupo)
    pend = sb_rest("GET", f"/group_join_requests?user_id=eq.{u['id']}&status=eq.pending&select=group_id")
    pend_nomes = []
    if pend:
        pids = ",".join(f'"{p["group_id"]}"' for p in pend)
        pend_nomes = [g["name"] for g in sb_rest("GET", f"/groups?id=in.({pids})&select=name")]
    if not memberships:
        return {"grupos": [], "nome": u["nome"], "pendentes": pend_nomes}
    ids = ",".join(f'"{m["group_id"]}"' for m in memberships)
    grupos = {g["id"]: g["name"] for g in sb_rest("GET", f"/groups?id=in.({ids})&select=id,name")}
    return {"nome": u["nome"],
            "grupos": [{"group_id": m["group_id"], "role": m["role"],
                        "nome": grupos.get(m["group_id"], m["group_id"][:8])} for m in memberships],
            "pendentes": pend_nomes}

@app.post("/api/grupos/solicitar")
def solicitar_entrada(c: CodigoAcesso, request: Request, bg: BackgroundTasks):
    """Solicita entrada em um grupo via código de acesso — aguarda aprovação do admin."""
    u = auth_jwt_only(request)
    codigo = c.codigo.strip().lower()
    if not codigo:
        raise HTTPException(400, "Informe o código de acesso.")
    grupos = sb_rest("GET", f"/groups?invite_code=ilike.{codigo}&select=id,name")
    if not grupos:
        raise HTTPException(404, "Código não encontrado. Confira com o administrador do grupo.")
    grupo = grupos[0]
    if any(m["group_id"] == grupo["id"] for m in get_memberships(u["id"])):
        raise HTTPException(400, f"Você já é membro do grupo {grupo['name']}.")
    pend = sb_rest("GET", f"/group_join_requests?user_id=eq.{u['id']}&group_id=eq.{grupo['id']}&status=eq.pending&select=id")
    if pend:
        raise HTTPException(400, "Você já tem uma solicitação pendente para esse grupo. Aguarde a aprovação.")
    # garante profile
    try:
        if not sb_rest("GET", f"/profiles?id=eq.{u['id']}&select=id"):
            sb_rest("POST", "/profiles", {"id": u["id"], "email": u.get("email", ""), "full_name": u["nome"]})
    except Exception as e:
        log.warning(f"profiles no solicitar: {e}")
    sb_rest("POST", "/group_join_requests",
            {"group_id": grupo["id"], "user_id": u["id"], "requested_role": "member", "status": "pending"})
    if VAPID_PUBLIC_KEY:
        bg.add_task(push_all_org, grupo["id"], "👤 Solicitação de entrada",
                    f"{u['nome']} pediu para entrar no grupo {grupo['name']}")
    log.info(f"Solicitação de entrada: {u['nome']} → {grupo['name']}")
    return {"ok": True, "grupo": grupo["name"],
            "mensagem": f"Solicitação enviada! Aguarde a aprovação do administrador de {grupo['name']}."}

@app.get("/api/grupos/solicitacoes")
def listar_solicitacoes(user=Depends(auth)):
    """Solicitações pendentes do grupo ativo (admin)."""
    if user["role"] != "admin":
        return []
    gid = user.get("org_id")
    rows = sb_rest("GET", f"/group_join_requests?group_id=eq.{gid}&status=eq.pending&select=*&order=created_at")
    if not rows:
        return []
    uids = ",".join(f'"{r["user_id"]}"' for r in rows)
    perfis = {p["id"]: p for p in sb_rest("GET", f"/profiles?id=in.({uids})&select=id,full_name,email")}
    return [{"id": r["id"],
             "nome": perfis.get(r["user_id"], {}).get("full_name") or "—",
             "email": perfis.get(r["user_id"], {}).get("email") or "",
             "role": r.get("requested_role", "member")} for r in rows]


def welcome_email_html(nome: str, grupo: str) -> str:
    """Email de boas-vindas — visual da marca, compatível com Gmail/Outlook (tabelas + estilos inline)."""
    base = (APP_BASE_URL or "").rstrip("/")
    logo = f"{base}/icon-192.png" if base else ""
    logo_html = (f'<img src="{logo}" width="76" height="76" alt="A.N.A" '
                 f'style="display:block;border-radius:18px">' if logo else
                 '<div style="font-size:30px;font-weight:700;color:#818cf8;letter-spacing:4px">A.N.A</div>')
    primeiro = (nome or "").split(" ")[0] or "colega"
    return f"""<!doctype html>
<html><body style="margin:0;padding:0;background:#f1f5f9;font-family:'Segoe UI',Arial,sans-serif">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f1f5f9;padding:28px 12px">
<tr><td align="center">
<table role="presentation" width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%;background:#ffffff;border-radius:16px;overflow:hidden;box-shadow:0 2px 12px rgba(15,23,42,.08)">

  <tr><td align="center" style="background:#0b1120;padding:34px 24px 26px">
    {logo_html}
    <div style="font-size:24px;font-weight:700;color:#eef2ff;letter-spacing:6px;margin-top:14px">A.N.A</div>
    <div style="font-size:12px;color:#67e8f9;letter-spacing:2px;margin-top:4px">SECRET&Aacute;RIA VIRTUAL DE ANESTESIOLOGIA</div>
  </td></tr>

  <tr><td style="padding:34px 36px 8px">
    <div style="font-size:19px;font-weight:700;color:#0f172a">Bem-vindo(a), Dr(a). {primeiro}! 🎉</div>
    <p style="font-size:14px;line-height:1.7;color:#334155;margin:14px 0 0">
      Sua entrada no grupo <b style="color:#4338ca">{grupo}</b> foi aprovada.
      A partir de agora voc&ecirc; tem acesso &agrave; <b>A.N.A</b> — a assistente com intelig&ecirc;ncia artificial
      que cuida da agenda e da escala do seu grupo de anestesiologia.
    </p>
  </td></tr>

  <tr><td style="padding:18px 36px 6px">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
      <tr><td style="padding:11px 14px;background:#f8fafc;border-radius:10px;border-left:3px solid #22d3ee">
        <div style="font-size:13px;font-weight:700;color:#0f172a">💬 Agendamento em linguagem natural</div>
        <div style="font-size:12px;color:#64748b;margin-top:3px">Converse com a A.N.A pelo chat — ou envie a <b>foto</b> ou o <b>PDF</b> do mapa cir&uacute;rgico e ela agenda todos os procedimentos de uma vez.</div>
      </td></tr>
      <tr><td style="height:8px"></td></tr>
      <tr><td style="padding:11px 14px;background:#f8fafc;border-radius:10px;border-left:3px solid #6366f1">
        <div style="font-size:13px;font-weight:700;color:#0f172a">🗓️ Escala de plant&otilde;es integrada</div>
        <div style="font-size:12px;color:#64748b;margin-top:3px">Veja e edite a escala do grupo, proponha <b>trocas</b>, anuncie plant&otilde;es para colegas assumirem e acompanhe seus <b>cr&eacute;ditos</b> do m&ecirc;s.</div>
      </td></tr>
      <tr><td style="height:8px"></td></tr>
      <tr><td style="padding:11px 14px;background:#f8fafc;border-radius:10px;border-left:3px solid #22d3ee">
        <div style="font-size:13px;font-weight:700;color:#0f172a">🔔 Avisos inteligentes</div>
        <div style="font-size:12px;color:#64748b;margin-top:3px">Notifica&ccedil;&otilde;es de novos agendamentos, trocas e o resumo da agenda do dia seguinte — com sincroniza&ccedil;&atilde;o opcional ao seu <b>Google Calendar</b>.</div>
      </td></tr>
    </table>
  </td></tr>

  <tr><td align="center" style="padding:26px 36px 8px">
    <a href="{base or '#'}" style="display:inline-block;background:#4338ca;color:#ffffff;text-decoration:none;font-size:14px;font-weight:700;padding:13px 34px;border-radius:10px">Acessar a A.N.A</a>
    <div style="font-size:11px;color:#94a3b8;margin-top:12px">💡 Dica: no celular, use <b>&ldquo;Adicionar &agrave; tela de in&iacute;cio&rdquo;</b> para instalar a A.N.A como aplicativo.</div>
  </td></tr>

  <tr><td style="padding:22px 36px 26px">
    <div style="border-top:1px solid #e2e8f0;padding-top:14px;font-size:11px;color:#94a3b8;text-align:center">
      Voc&ecirc; recebeu este email porque sua entrada no grupo {grupo} foi aprovada pelo administrador.<br>
      A.N.A &middot; Secret&aacute;ria Virtual de Anestesiologia
    </div>
  </td></tr>
</table>
</td></tr></table>
</body></html>"""

@app.post("/api/grupos/solicitacoes/{req_id}/responder")
def responder_solicitacao(req_id: str, r: SolicitacaoResposta, bg: BackgroundTasks, user=Depends(auth)):
    if user["role"] != "admin":
        raise HTTPException(403, "Apenas administradores podem aprovar solicitações.")
    gid = user.get("org_id")
    rows = sb_rest("GET", f"/group_join_requests?id=eq.{req_id}&group_id=eq.{gid}&select=*")
    if not rows:
        raise HTTPException(404, "Solicitação não encontrada.")
    req = rows[0]
    if req.get("status") != "pending":
        raise HTTPException(400, "Solicitação já respondida.")
    if r.aprovar:
        sb_rest("POST", "/group_members",
                {"user_id": req["user_id"], "group_id": gid, "role": req.get("requested_role", "member")})
        _membership_cache.pop(req["user_id"], None)
        # email de boas-vindas ao novo integrante
        try:
            perfil = sb_rest("GET", f"/profiles?id=eq.{req['user_id']}&select=full_name,email")
            grupo = sb_rest("GET", f"/groups?id=eq.{gid}&select=name")
            if perfil and perfil[0].get("email"):
                bg.add_task(send_email, perfil[0]["email"],
                            f"🎉 Bem-vindo(a) ao grupo {grupo[0]['name'] if grupo else ''} — A.N.A",
                            welcome_email_html(perfil[0].get("full_name") or "", grupo[0]["name"] if grupo else "seu grupo"))
        except Exception as e:
            log.warning(f"email boas-vindas: {e}")
    sb_rest("PATCH", f"/group_join_requests?id=eq.{req_id}",
            {"status": "approved" if r.aprovar else "rejected"})
    log.info(f"Solicitação {'aprovada' if r.aprovar else 'recusada'}: {req_id}")
    return {"ok": True}

@app.get("/api/grupos/membros")
def listar_membros(user=Depends(auth)):
    """Integrantes do grupo ativo, com perfil e médico vinculado."""
    gid = user.get("org_id")
    rows = sb_rest("GET", f"/group_members?group_id=eq.{gid}&select=user_id,role,created_at&order=created_at")
    if not rows:
        return []
    uids = ",".join(f'"{r["user_id"]}"' for r in rows)
    perfis = {p["id"]: p for p in sb_rest("GET", f"/profiles?id=in.({uids})&select=id,full_name,email")}
    docs = {d["user_id"]: d["name"] for d in
            sb_rest("GET", f"/doctors?group_id=eq.{gid}&user_id=not.is.null&select=name,user_id")}
    return [{"user_id": r["user_id"],
             "nome": perfis.get(r["user_id"], {}).get("full_name") or "—",
             "email": perfis.get(r["user_id"], {}).get("email") or "",
             "role": r["role"],
             "medico": docs.get(r["user_id"]),
             "sou_eu": r["user_id"] == user["id"]} for r in rows]

@app.delete("/api/grupos/membros/{uid}")
def remover_membro(uid: str, user=Depends(auth)):
    """Remove um integrante do grupo (admin). O acesso dele à Ana e ao MedSS neste grupo é revogado."""
    if user["role"] != "admin":
        raise HTTPException(403, "Apenas administradores podem remover integrantes.")
    if uid == user["id"]:
        raise HTTPException(400, "Você não pode remover a si mesmo do grupo.")
    gid = user.get("org_id")
    rows = sb_rest("GET", f"/group_members?group_id=eq.{gid}&user_id=eq.{uid}&select=id")
    if not rows:
        raise HTTPException(404, "Integrante não encontrado neste grupo.")
    sb_rest("DELETE", f"/group_members?group_id=eq.{gid}&user_id=eq.{uid}")
    # desativa o médico vinculado — sai da escala e das listas (histórico preservado)
    try:
        sb_rest("PATCH", f"/doctors?group_id=eq.{gid}&user_id=eq.{uid}", {"active": False})
    except Exception as e:
        log.warning(f"desativar médico do removido: {e}")
    _membership_cache.pop(uid, None)
    log.info(f"Membro removido do grupo {gid}: {uid}")
    return {"ok": True}

@app.get("/api/grupos/codigo")
def codigo_acesso(user=Depends(auth)):
    """Código de acesso do grupo ativo (admin) — para compartilhar."""
    if user["role"] != "admin":
        raise HTTPException(403, "Apenas administradores.")
    rows = sb_rest("GET", f"/groups?id=eq.{user.get('org_id')}&select=invite_code")
    return {"codigo": (rows[0].get("invite_code") or "").upper() if rows else ""}

@app.get("/api/supabase/status")
def supabase_status():
    """Diagnóstico da integração Supabase (sem expor chaves)."""
    status = {"configured": bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY),
              "url_set": bool(SUPABASE_URL),
              "anon_key_set": bool(SUPABASE_ANON_KEY),
              "service_role_set": bool(SUPABASE_SERVICE_ROLE_KEY),
              "jwks_reachable": False, "rest_reachable": False}
    if SUPABASE_URL:
        try:
            req = urllib.request.Request(f"{SUPABASE_URL}/auth/v1/.well-known/jwks.json")
            with urllib.request.urlopen(req, timeout=8) as r:
                jwks = json.loads(r.read())
                status["jwks_reachable"] = bool(jwks.get("keys"))
        except Exception as e:
            status["jwks_error"] = str(e)[:150]
        if SUPABASE_SERVICE_ROLE_KEY:
            try:
                rows = sb_rest("GET", "/groups?select=id&limit=1")
                status["rest_reachable"] = True
                status["groups_visible"] = len(rows)
            except Exception as e:
                status["rest_error"] = str(e)[:150]
    return status

def auth_jwt_only(request: Request) -> dict:
    """Valida apenas o JWT (sem exigir grupo) — para onboarding e whoami."""
    authz = request.headers.get("Authorization", "")
    if not (authz.startswith("Bearer ") and SUPABASE_URL):
        raise HTTPException(401, "Não autorizado.")
    claims = validate_supabase_jwt(authz[7:])
    if not claims or not claims.get("sub"):
        raise HTTPException(401, "Não autorizado.")
    meta = claims.get("user_metadata") or {}
    email = claims.get("email", "")
    nome = meta.get("full_name") or meta.get("name") or (email.split("@")[0] if email else claims["sub"][:8])
    return {"id": claims["sub"], "nome": nome, "email": email}

class NovoGrupo(BaseModel):
    nome: str

@app.post("/api/grupos")
def criar_grupo(g: NovoGrupo, request: Request):
    """Cria um grupo novo e torna o usuário atual admin (onboarding)."""
    u = auth_jwt_only(request)
    nome = g.nome.strip()[:120]
    if not nome:
        raise HTTPException(400, "Informe o nome do grupo.")
    # garante profile (FK de group_members/created_by pode exigir)
    try:
        existing = sb_rest("GET", f"/profiles?id=eq.{u['id']}&select=id")
        if not existing:
            sb_rest("POST", "/profiles", {"id": u["id"], "full_name": u["nome"]})
    except Exception as e:
        log.warning(f"profiles check/insert: {e}")
    grupo = sb_rest("POST", "/groups", {"name": nome, "created_by": u["id"]})[0]
    sb_rest("POST", "/group_members",
            {"user_id": u["id"], "group_id": grupo["id"], "role": "admin"})
    _membership_cache.pop(u["id"], None)
    log.info(f"Grupo criado no onboarding: {nome} por {u['nome']}")
    return {"ok": True, "group_id": grupo["id"], "nome": nome}

@app.get("/api/supabase/whoami")
def supabase_whoami(request: Request):
    """Identidade + grupos — funciona mesmo sem grupo (para onboarding)."""
    u = auth_jwt_only(request)
    memberships = get_memberships(u["id"])
    wanted = request.headers.get("X-Group-Id", "")
    m = None
    if memberships:
        m = next((x for x in memberships if x["group_id"] == wanted), None) or memberships[0]
    return {"id": u["id"], "nome": u["nome"],
            "role": ("admin" if m and m.get("role") == "admin" else "medico") if m else None,
            "group_id": m["group_id"] if m else None,
            "auth_source": "supabase", "memberships": memberships}

@app.get("/api/health")
def health():
    return {"status":"ok","version":"3.0.0",
            "db":"supabase",
            "email":bool(SMTP_HOST),"gcal":bool(GCAL_CREDS),
            "ai":bool(GROQ_KEY),
            "oauth_google":bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET and APP_BASE_URL),
            "google_routes":bool(GOOGLE_ROUTES_API_KEY),
            "gemini":bool(GEMINI_API_KEY),
            "ocr_space":bool(OCR_SPACE_API_KEY),
            "supabase":bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY),
            "timestamp":datetime.now().isoformat()}

# ── PÁGINAS PÚBLICAS (política de privacidade / termos) ────
PRIVACY_HTML = """<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Política de Privacidade — Ana Secretária Virtual</title>
<style>
body{font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;max-width:700px;margin:0 auto;padding:32px 20px;color:#1a1a2e;line-height:1.6}
h1{font-size:22px;color:#6C63D4;margin-bottom:4px}
h2{font-size:16px;color:#3C3489;margin-top:28px}
p,li{font-size:14px;color:#333}
.upd{font-size:12px;color:#888;margin-bottom:24px}
a{color:#6C63D4}
ul{padding-left:20px}
</style></head>
<body>
<h1>Política de Privacidade</h1>
<div class="upd">Ana — Secretária Virtual de Anestesiologia · Última atualização: junho de 2026</div>

<p>Esta política descreve como a Ana, sistema de agendamento para grupos de anestesiologia, coleta, usa e protege as informações dos usuários.</p>

<h2>Quem somos</h2>
<p>A Ana é uma aplicação de uso interno desenvolvida para auxiliar grupos de profissionais de anestesiologia a organizar agendamentos de procedimentos médicos. Contato: <a href="mailto:rodrigomorlin@gmail.com">rodrigomorlin@gmail.com</a>.</p>

<h2>Quais dados coletamos</h2>
<ul>
<li><b>Dados de cadastro:</b> nome, identificador de usuário e PIN de acesso (armazenado de forma criptografada).</li>
<li><b>Dados de agendamento:</b> data, horário, setor, procedimento, médico responsável e, quando informado, nome do paciente.</li>
<li><b>Documentos anexados:</b> PDFs de pedidos médicos enviados pelo usuário para agendamento automático.</li>
<li><b>Dados do Google Calendar:</b> quando o usuário conecta sua conta Google, criamos e gerenciamos eventos no calendário pessoal dele, exclusivamente para refletir os agendamentos feitos dentro da Ana.</li>
</ul>

<h2>Como usamos os dados</h2>
<p>Os dados são usados exclusivamente para: organizar a agenda de procedimentos do grupo, sincronizar agendamentos com o Google Calendar do usuário (quando autorizado), enviar notificações por email sobre agendamentos, e gerar relatórios internos de uso do grupo.</p>

<h2>Acesso ao Google Calendar</h2>
<p>Ao conectar sua conta Google, a Ana solicita permissão para criar, editar e excluir eventos no seu calendário. Essa permissão é usada apenas para refletir os agendamentos feitos através do sistema. A Ana não lê, acessa ou compartilha outros eventos já existentes no seu calendário além dos que ela mesma cria.</p>

<h2>Compartilhamento de dados</h2>
<p>Os dados não são vendidos, alugados ou compartilhados com terceiros para fins de marketing. Dados podem ser processados por provedores de infraestrutura (hospedagem em nuvem) e de inteligência artificial (para interpretação de linguagem natural dos pedidos de agendamento), estritamente para o funcionamento do serviço.</p>

<h2>Retenção e exclusão de dados</h2>
<p>Os dados são mantidos enquanto a conta do usuário estiver ativa. O usuário pode solicitar a exclusão de seus dados e a desconexão do Google Calendar a qualquer momento, através do administrador do sistema ou diretamente na tela de configurações (opção "Desconectar").</p>

<h2>Segurança</h2>
<p>PINs de acesso são armazenados com hash criptográfico. Tokens de acesso ao Google são armazenados de forma segura e usados apenas para as operações descritas nesta política.</p>

<h2>Alterações nesta política</h2>
<p>Esta política pode ser atualizada periodicamente. A data da última atualização está sempre indicada no topo desta página.</p>

<h2>Contato</h2>
<p>Para dúvidas sobre esta política ou para solicitar a exclusão de dados, entre em contato: <a href="mailto:rodrigomorlin@gmail.com">rodrigomorlin@gmail.com</a>.</p>

</body></html>"""

TERMS_HTML = """<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Termos de Serviço — Ana Secretária Virtual</title>
<style>
body{font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;max-width:700px;margin:0 auto;padding:32px 20px;color:#1a1a2e;line-height:1.6}
h1{font-size:22px;color:#6C63D4;margin-bottom:4px}
h2{font-size:16px;color:#3C3489;margin-top:28px}
p,li{font-size:14px;color:#333}
.upd{font-size:12px;color:#888;margin-bottom:24px}
a{color:#6C63D4}
</style></head>
<body>
<h1>Termos de Serviço</h1>
<div class="upd">Ana — Secretária Virtual de Anestesiologia · Última atualização: junho de 2026</div>

<p>Ao utilizar a Ana, você concorda com os termos abaixo.</p>

<h2>Uso do serviço</h2>
<p>A Ana é destinada ao uso interno de grupos de profissionais de anestesiologia para fins de organização de agendamentos. O acesso é restrito a usuários autorizados pelo administrador do grupo.</p>

<h2>Responsabilidade pelos dados inseridos</h2>
<p>O usuário é responsável pela exatidão das informações inseridas no sistema, incluindo dados de agendamento e documentos anexados. A Ana é uma ferramenta de apoio organizacional e não substitui o julgamento clínico profissional.</p>

<h2>Integração com Google Calendar</h2>
<p>A conexão com o Google Calendar é opcional e pode ser desfeita pelo usuário a qualquer momento na tela de configurações. Ao conectar, o usuário autoriza a Ana a criar, editar e remover eventos correspondentes aos agendamentos feitos no sistema.</p>

<h2>Disponibilidade</h2>
<p>O serviço é fornecido "como está", sem garantias de disponibilidade contínua. Esforços razoáveis são feitos para manter o sistema funcionando, mas interrupções podem ocorrer.</p>

<h2>Alterações</h2>
<p>Estes termos podem ser atualizados periodicamente, com a data de revisão indicada no topo desta página.</p>

<h2>Contato</h2>
<p><a href="mailto:rodrigomorlin@gmail.com">rodrigomorlin@gmail.com</a></p>

</body></html>"""

@app.get("/privacidade", response_class=HTMLResponse)
def privacy_policy():
    return HTMLResponse(PRIVACY_HTML)

@app.get("/termos", response_class=HTMLResponse)
def terms_of_service():
    return HTMLResponse(TERMS_HTML)

# ── STATIC FILES ───────────────────────────────────────────
@app.get("/sw.js")
def sw(): return HTMLResponse(open("sw.js").read(), media_type="application/javascript")

@app.get("/manifest.json")
def manifest(): return JSONResponse(json.load(open("manifest.json")))

from fastapi.responses import FileResponse as _FileResponse
import os as _os
_APP_DIR = _os.path.dirname(_os.path.abspath(__file__))

@app.get("/api/geocode")
def geocode(q: str, user=Depends(auth)):
    """Busca endereço pelo nome do local via Google Places Text Search (mesma conta da chave de rotas)."""
    if not GOOGLE_ROUTES_API_KEY:
        raise HTTPException(400, "GOOGLE_ROUTES_API_KEY não configurada.")
    if not q or len(q.strip()) < 3:
        raise HTTPException(400, "Digite pelo menos 3 caracteres.")
    payload = json.dumps({"textQuery": q.strip(), "languageCode": "pt-BR", "maxResultCount": 4}).encode("utf-8")
    http_req = urllib.request.Request(
        "https://places.googleapis.com/v1/places:searchText",
        data=payload, method="POST",
        headers={"Content-Type": "application/json",
                 "X-Goog-Api-Key": GOOGLE_ROUTES_API_KEY,
                 "X-Goog-FieldMask": "places.displayName,places.formattedAddress"})
    try:
        with urllib.request.urlopen(http_req, timeout=15) as resp:
            j = json.loads(resp.read())
        out = [{"nome": p.get("displayName", {}).get("text", ""),
                "endereco": p.get("formattedAddress", "")}
               for p in j.get("places", []) if p.get("formattedAddress")]
        return {"resultados": out}
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        log.error(f"Places search {e.code}: {body[:300]}")
        if e.code == 403:
            raise HTTPException(502, "A 'Places API (New)' não está habilitada nesta chave do Google. "
                                     "Habilite em console.cloud.google.com → APIs & Services → Places API (New) → Enable.")
        raise HTTPException(502, f"Erro na busca de endereço ({e.code}).")

@app.get("/icon-192.png")
def icon_192():
    return _FileResponse(_os.path.join(_APP_DIR, "icon-192.png"), media_type="image/png",
                         headers={"Cache-Control": "public, max-age=86400"})

@app.get("/icon-512.png")
def icon_512():
    return _FileResponse(_os.path.join(_APP_DIR, "icon-512.png"), media_type="image/png",
                         headers={"Cache-Control": "public, max-age=86400"})

@app.get("/icon-maskable-512.png")
def icon_maskable():
    return _FileResponse(_os.path.join(_APP_DIR, "icon-maskable-512.png"), media_type="image/png",
                         headers={"Cache-Control": "public, max-age=86400"})

@app.get("/apple-touch-icon.png")
def apple_touch_icon():
    return _FileResponse(_os.path.join(_APP_DIR, "apple-touch-icon.png"), media_type="image/png",
                         headers={"Cache-Control": "public, max-age=86400"})

@app.get("/icon.svg")
def icon_svg():
    svg = '''<svg viewBox="0 0 512 512" xmlns="http://www.w3.org/2000/svg">
<rect width="512" height="512" rx="96" fill="#0b1120"/>
<g transform="translate(256 256)">
<circle r="205" fill="none" stroke="#67e8f9" stroke-width="11" opacity=".85" stroke-dasharray="1.8 9.2"/>
<circle r="168" fill="none" stroke="#22d3ee" stroke-width="9" stroke-linecap="round" stroke-dasharray="216 136.4"/>
<circle r="144" fill="none" stroke="#22d3ee" stroke-width="14" opacity=".5" stroke-dasharray="3.6 4.8 9.6 4.8 3.6 10.8 16.8 6"/>
<circle r="101" fill="none" stroke="#22d3ee" stroke-width="5"/>
<circle r="92" fill="#22d3ee" opacity=".08"/>
<text x="4" y="31" text-anchor="middle" font-family="Arial,sans-serif" font-size="91" font-weight="700" letter-spacing="7" fill="#818cf8">A.N.A</text>
</g>
</svg>'''
    return Response(content=svg, media_type="image/svg+xml",
                    headers={"Cache-Control": "public, max-age=86400"})
@app.get("/", response_class=HTMLResponse)
def index(): return HTMLResponse(open("index.html", encoding="utf-8").read())

# ── Inicializa a camada Supabase (feature flag ANA_DATA_BACKEND) ──
ana_data.init(sb_rest=sb_rest, log=log,
              routes_duration=_google_routes_duration,
              default_group="")
log.info("A.N.A · backend Supabase (banco compartilhado com o MedSS)")
